#TODO: File location naming
#TODO: Sheet naming

import os
import openpyxl
import re
from datetime import date, datetime, timedelta

os.chdir('C:\\Users\yasirc\Desktop\parseData') # enter correct filepath for project here
wb = openpyxl.load_workbook('ngt_log.xlsx') # filename being analyzed
sheet = wb['sheet1'] # name of the sheet that is being analyzed
cycleTimes = []
coulCount = []
capChange = []
dateTimes = []
timeDeltas = []

cycleTimes.append(sheet.cell(row=2, column=2).value.time()) # adds 00:00:00 as start time
coulCount.append(sheet.cell(row=2, column=8).value)

print("ROW | VALUE")
for i in range(2, sheet.max_row):
    current = sheet.cell(row=i, column=6).value
    next =  sheet.cell(row=i+1, column=6).value
    if int(current) >= 0 and int(next) <= 0 or int(current) <= 0 and int(next) >= 0:
        cycleTimes.append(sheet.cell(row=i, column=2).value)
        cycleTimes.append(sheet.cell(row=(i+1), column=2).value)
        coulCount.append(sheet.cell(row=i, column=8).value)
        coulCount.append(sheet.cell(row=(i+1), column=8).value)
        print(i+1, next)

print("\nCYCLE CHANGE TIMES")        
cycleTimes.append(sheet.cell(row=sheet.max_row, column=2).value)        
for i in cycleTimes:
    print(i)
    dateTimes.append(datetime.combine(date.today(), i))

print("\nELAPSED TIMES PER CYCLE")
starts = dateTimes[::2]
ends = dateTimes[1::2]
timeDeltas = [end - start for start, end in zip(starts, ends)]
for i in timeDeltas:
    print(i)

print("\nCOLOUMB COUNT")
coulCount.append(sheet.cell(row=sheet.max_row, column=8).value)
for i in coulCount:
    print(i)

print("\nCAPACITY CHANGE PER CYCLE")
starts = coulCount[::2]
ends = coulCount[1::2]
capChange = [end - start for start, end in zip(starts, ends)]
for i in capChange:
    print(i)

sheet['B2'] = sheet.cell(row=2, column=2).value.time()

wb.create_sheet('sheet2') # insert at the end (default)
ws1 = wb.active
ws2 = wb['sheet2']
for cell in ws1['B:B']:
#    print('Printing from column ' + str(cell.column) + ' row ' + str(cell.row))
    ws2.cell(row = cell.row, column = 1, value = cell.value)

for cell in ws1['D:D']:
#    print('Printing from column ' + str(cell.column) + ' row ' + str(cell.row))
    ws2.cell(row = cell.row, column = 2, value = cell.value)   

print('Creating charts...')

sheet = wb['sheet2'] # focus on sheet2 to pull data from/write chart to
refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=3, max_col=2, max_row=sheet.max_row)
seriesObj = openpyxl.chart.Series(refObj, title='Series1')
chartObj = openpyxl.chart.LineChart()
chartObj.title = 'SLA Discharge - 5.5A: V_BAT'
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'C5')
wb.save('ngt_log.xlsx')



# input() #keeps cmd window open after script execution
