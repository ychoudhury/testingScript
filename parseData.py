#TODO: File location naming
#TODO: Sheet naming

import os
import re
import openpyxl
from openpyxl import Workbook, chart
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.axis import DateAxis
from datetime import date, datetime, timedelta, time

# enter correct filepath for project here
os.chdir('C:\\Users\yasirc\Desktop\parseData')

# enter correct filename here
wb = openpyxl.load_workbook('ngt_log.xlsx')

# name of the sheet that is being analyzed
sheet = wb['sheet1']

cycleTimes = []
coulCount = []
capChange = []
dateTimes = []
timeDeltas = []
graphIntervals = []

# RegEx to convert times in column B to a consistent format
def cell_to_datetime(cell):
    if isinstance(cell.value, datetime):
        return datetime.combine(date.today(), cell.value.time())

    match = re.search('(\d+) day[s]?, (\d+):(\d+):(\d+)', str(cell.value))

    if not match:
        match = re.search('(\d+):(\d+):(\d+)', str(cell.value))

        if match:
            cell_time = time(int(match.group(1)), int(match.group(2)), int(match.group(3)))
            return datetime.combine(date.today(), cell_time)
        else:
            raise Exception('no match for time value')

    cell_time = time(int(match.group(2)), int(match.group(3)), int(match.group(4)))
    cell_datetime = datetime.combine(date.today(), cell_time)
    adjusted_datetime = cell_datetime + timedelta(days=int(match.group(1)))

    return adjusted_datetime

# first value added to array for time calculation purposes
cycleTimes.append(cell_to_datetime(sheet.cell(row=2, column=2)))
coulCount.append(sheet.cell(row=2, column=8).value)
graphIntervals.append(sheet.cell(row=2, column=1).value)

# finds when battery current changes from + to -/- to +
for i in range(2, sheet.max_row):
    current = sheet.cell(row=i, column=6).value
    next =  sheet.cell(row=i+1, column=6).value
    if int(current) >= 0 and int(next) <= 0 or int(current) <= 0 and int(next) >= 0:
        cycleTimes.append(cell_to_datetime(sheet.cell(row=i, column=2)))
        cycleTimes.append(cell_to_datetime(sheet.cell(row=(i+1), column=2)))
        coulCount.append(sheet.cell(row=i, column=8).value)
        coulCount.append(sheet.cell(row=(i+1), column=8).value)
        graphIntervals.append(sheet.cell(row=i, column=1).value)
        graphIntervals.append(sheet.cell(row=(i+1), column=1).value)

print("\nCYCLE CHANGE TIMES")        
cycleTimes.append(cell_to_datetime(sheet.cell(row=sheet.max_row, column=2)))
for i in cycleTimes:
    print(i)

print("\nELAPSED TIMES PER CYCLE")
starts = cycleTimes[::2]
ends = cycleTimes[1::2]
timeDeltas = [end - start for start, end in zip(starts, ends)]
for i in timeDeltas:
    print(i)

print("\nCOLOUMB COUNT")
coulCount.append(sheet.cell(row=sheet.max_row, column=8).value)
for i in coulCount:
    print(i)

print("\nCAPACITY CHANGE PER CYCLE")
starts = coulCount[::2]
ends = coulCount[1::2]
capChange = [end - start for start, end in zip(starts, ends)]
for i in capChange:
    print(i)

graphIntervals.append(sheet.cell(row=sheet.max_row, column=1).value)

# chart creation
wb.create_sheet('Data Analysis') # new sheet created for data analysis
ws1 = wb['sheet1']
ws2 = wb['Data Analysis']
for cell in ws1['B:B']:
    ws2.cell(row = cell.row, column = 1, value = cell.value)
for cell in ws1['D:D']:
    ws2.cell(row = cell.row, column = 2, value = cell.value)
for cell in ws1['I:I']:
    ws2.cell(row = cell.row, column = 3, value = cell.value)

print("\nCreating charts...")
sheet = wb['Data Analysis'] # focus on Data Analysis sheet to pull data from/write chart to    
for i in range(2, sheet.max_row):
    cell = sheet.cell(row=i, column=1)
    cell.value = str(cell_to_datetime(cell))

dates = chart.Reference(ws2, min_col=1, min_row=2, max_row=sheet.max_row)
vBat = chart.Reference(ws2, min_col=2, min_row=1, max_col=2, max_row=sheet.max_row)
qBat = chart.Reference(ws2, min_col=3, min_row=1, max_col=3, max_row=sheet.max_row)
c1 = chart.LineChart()
c1.title = "SLA Discharge - 5.5A: V_BAT and Q_Count"
c1.height = 10 # chart dimensions
c1.width = 20 
c1.x_axis.majorTimeUnit = "days"
c1.x_axis = chart.axis.DateAxis()
c1.x_axis.title = "Time"
c1.x_axis.crosses = "min"
c1.x_axis.majorTickMark = "out"
c1.x_axis.number_format = 'd-HH-MM-SS'
c1.add_data(vBat, titles_from_data=True)
c1.set_categories(dates)
c1.y_axis.title = "Battery Voltage"
c1.y_axis.scaling.min = 10
c1.y_axis.scaling.max = 14500
c1.y_axis.crossAx = 500
c1.y_axis.majorGridlines = None

c2 = chart.LineChart()
c2.height = 15
c2.width = 25
c2.x_axis.axId = 500
c2.add_data(qBat, titles_from_data=True)
c2.y_axis.axId = 200
c2.y_axis.title = "Qbat Percentage"
c2.y_axis.crossAx = 500

c1.y_axis.crosses = "max"
c1 += c2

# Line styling
s1 = c1.series[0]
s1.graphicalProperties.line.solidFill = "BE4B48"
s1.graphicalProperties.line.width = 25000
s1.smooth = True # make the line smooth
s2 = c2.series[0]
s2.graphicalProperties.line.solidFill = "4A7EBB"
s2.graphicalProperties.line.width = 25000
s2.smooth = True # make the line smooth


ws2.add_chart(c1, "P5") # positioning of large chart

chart_row = 5

for i in range(0, len(graphIntervals), 2):
	min_row = graphIntervals[i] + 1
	max_row = graphIntervals[i+1] + 1

	# skip headers on first row
	if min_row == 1:
		min_row = 2

	dates = chart.Reference(ws2, min_col=1, min_row=min_row, max_row=max_row)
	vBat = chart.Reference(ws2, min_col=2, min_row=min_row, max_col=2, max_row=max_row)
	qBat = chart.Reference(ws2, min_col=3, min_row=min_row, max_col=3, max_row=max_row)

	c1 = chart.LineChart()
	c1.title = "SLA Discharge - 5.5A: V_BAT and Q_Count"
	c1.x_axis.majorTimeUnit = "days"
	c1.x_axis = chart.axis.DateAxis()
	c1.x_axis.title = "Time"
	c1.x_axis.crosses = "min"
	c1.x_axis.majorTickMark = "out"
	c1.x_axis.number_format = 'd-HH-MM-SS'
	c1.append(Series(vBat, title="Battery Voltage"))
	c1.set_categories(dates)
	c1.y_axis.title = "Battery Voltage"
	c1.height = 15
	c1.width = 20
	c1.y_axis.scaling.min = 10
	c1.y_axis.scaling.max = 14500
	c1.y_axis.crossAx = 500
	c1.y_axis.majorGridlines = None

	c2 = chart.LineChart()
	c2.height = 15
	c2.width = 20
	c2.x_axis.axId = 500
	c2.append(Series(qBat, title="Qbat Percentage"))
	c2.set_categories(dates)
	c2.y_axis.axId = 200
	c2.y_axis.title = "Qbat Percentage"
	c2.y_axis.crossAx = 500

	c1.y_axis.crosses = "max"
	c1 += c2

	s1 = c1.series[0]
	s1.graphicalProperties.line.solidFill = "BE4B48"
	s1.graphicalProperties.line.width = 25000
	s1.smooth = True # Make the line smooth
	s2 = c2.series[0]
	s2.graphicalProperties.line.solidFill = "4A7EBB"
	s2.graphicalProperties.line.width = 25000
	s2.smooth = True # Make the line smooth
	ws2.add_chart(c1, "D%d" % (chart_row))

	chart_row += 30

wb.save('ngt_log.xlsx')
