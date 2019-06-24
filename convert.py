import os
import openpyxl
import re
from datetime import date, datetime, timedelta, time

os.chdir('C:\\Users\Yasir\Desktop\parseData')

def cell_to_datetime(cell):
    if isinstance(cell.value, datetime):
        return datetime.combine(date.today(), cell.value.time())

    match = re.search('(\d+) day[s]?, (\d+):(\d+):(\d+)', str(cell.value))

    if not match:
        match = re.search('(\d+):(\d+):(\d+)', str(cell.value))

        if match:
            cell_time = time(int(match.group(1)), int(match.group(2)), int(match.group(3)))
            return datetime.combine(date.today(), cell_time)
        else:
            raise Exception('no match for time value')

    cell_time = time(int(match.group(2)), int(match.group(3)), int(match.group(4)))
    cell_datetime = datetime.combine(date.today(), cell_time)
    adjusted_datetime = cell_datetime + timedelta(days=int(match.group(1)))

    return adjusted_datetime

os.chdir('C:\\Users\Yasir\Desktop\parseData')
wb = openpyxl.load_workbook('ngt_log.xlsx')
sheet = wb['sheet1']

print('converting...')
for i in range(2, sheet.max_row):
    cell = sheet.cell(row=i, column=2)
    cell.value = str(cell_to_datetime(cell))

wb.create_sheet('sheet2') # insert at the end (default)
ws1 = wb['sheet1']
ws2 = wb['sheet2']
for cell in ws1['B:B']:
#    print('Printing from column ' + str(cell.column) + ' row ' + str(cell.row))
    ws2.cell(row = cell.row, column = 1, value = cell.value)
for cell in ws1['D:D']:
#    print('Printing from column ' + str(cell.column) + ' row ' + str(cell.row))
    ws2.cell(row = cell.row, column = 2, value = cell.value)   
print('Creating charts...')
sheet = wb['sheet2'] # focus on sheet2 to pull data from/write chart to
refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=2, max_row=sheet.max_row)
seriesObj = openpyxl.chart.Series(refObj, title='Series1')
chartObj = openpyxl.chart.LineChart()
chartObj.title = 'SLA Discharge - 5.5A: V_BAT'
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'C5')
wb.save('ngt_log.xlsx')

#print('saving...')
#wb.save('ngt_log.xlsx')
