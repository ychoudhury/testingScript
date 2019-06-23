#TODO: File location naming
#TODO: Sheet naming

import os
import openpyxl
import re
from datetime import date, datetime, timedelta, time

# enter correct filepath for project here
os.chdir('C:\\Users\Yasir\Desktop\parseData')

# enter correct filename here
wb = openpyxl.load_workbook('ngt_log.xlsx')

# name of the sheet that is being analyzed
sheet = wb['sheet1']

cycleTimes = []
coulCount = []
capChange = []
dateTimes = []
timeDeltas = []

def cell_to_datetime(cell):
    if isinstance(cell.value, datetime):
        return datetime.combine(date.today(), cell.value.time())

    match = re.search('(\d+) day[s]?, (\d+):(\d+):(\d+)', str(cell.value))

    if not match:
        match = re.search('(\d+):(\d+):(\d+)', str(cell.value))

        if match:
            cell_time = time(int(match.group(1)), int(match.group(2)), int(match.group(3)))
            return datetime.combine(date.today(), cell_time)
        else:
            raise Exception('no match for time value')

    cell_time = time(int(match.group(2)), int(match.group(3)), int(match.group(4)))
    cell_datetime = datetime.combine(date.today(), cell_time)
    adjusted_datetime = cell_datetime + timedelta(days=int(match.group(1)))

    return adjusted_datetime

# first value added to array for time calculation purposes.
cycleTimes.append(cell_to_datetime(sheet.cell(row=2, column=2)))
coulCount.append(sheet.cell(row=2, column=8).value)

# finds when battery current changes from + to -/- to +
print("ROW | VALUE")
for i in range(2, sheet.max_row):
    current = sheet.cell(row=i, column=6).value
    next =  sheet.cell(row=i+1, column=6).value
    if int(current) >= 0 and int(next) <= 0 or int(current) <= 0 and int(next) >= 0:
        cycleTimes.append(cell_to_datetime(sheet.cell(row=i, column=2)))
        cycleTimes.append(cell_to_datetime(sheet.cell(row=(i+1), column=2)))
        coulCount.append(sheet.cell(row=i, column=8).value)
        coulCount.append(sheet.cell(row=(i+1), column=8).value)
        print(i+1, next)

print("\nCYCLE CHANGE TIMES")        
cycleTimes.append(cell_to_datetime(sheet.cell(row=sheet.max_row, column=2)))
for i in cycleTimes:
    print(i)

print("\nELAPSED TIMES PER CYCLE")
starts = cycleTimes[::2]
ends = cycleTimes[1::2]
timeDeltas = [end - start for start, end in zip(starts, ends)]
for i in timeDeltas:
    print(i)

print("\nCOLOUMB COUNT")
coulCount.append(sheet.cell(row=sheet.max_row, column=8).value)
for i in coulCount:
    print(i)

print("\nCAPACITY CHANGE PER CYCLE")
starts = coulCount[::2]
ends = coulCount[1::2]
capChange = [end - start for start, end in zip(starts, ends)]
for i in capChange:
    print(i)

# sheet['B2'] = sheet.cell(row=2, column=2).value.time()

wb.save('ngt_log.xlsx')
